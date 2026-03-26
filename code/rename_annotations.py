"""
Copy annotation XMLs from the archive and rename annotation groups
based on the mapping defined in the Excel sheet.

Usage (on host):
    python3 code/rename_annotations.py

Usage (inside Docker container):
    python3 /home/user/source/code/rename_annotations.py

The script:
  1. Copies {slide_id}.xml files listed in the Excel from SOURCE_DIR to DEST_DIR.
  2. For every (slide_id, old_name → new_name) row, renames the annotation group
     in both <Annotation PartOfGroup="..."> and <Group Name="..."> elements.
"""

import shutil
import xml.etree.ElementTree as ET
from collections import defaultdict
from pathlib import Path

import openpyxl

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

SOURCE_DIR = Path(
    "/Volumes/PA_CPGARCHIVE/archives/kidney"
    "/wilmstumor_retrospective_prinsesmaximacentrum/annotations"
)
DEST_DIR = Path("/home/user/annotation")
EXCEL_PATH = Path(__file__).parent.parent / "Changes to annotations.xlsx"

# ---------------------------------------------------------------------------


def load_rename_map(excel_path: Path) -> dict[str, dict[str, str]]:
    """Return {slide_id: {old_name: new_name}} from the Excel sheet."""
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    slide_col = headers.index("Slide_id")
    old_col = headers.index("Old_Annotation_Name")
    new_col = headers.index("New_annotation_name")

    rename_map: dict[str, dict[str, str]] = defaultdict(dict)
    for row in ws.iter_rows(min_row=2, values_only=True):
        slide_id, old_name, new_name = row[slide_col], row[old_col], row[new_col]
        if slide_id and old_name and new_name:
            rename_map[str(slide_id)][str(old_name)] = str(new_name)

    return dict(rename_map)


def copy_annotations(slide_ids: list[str], source_dir: Path, dest_dir: Path) -> None:
    """Copy XMLs for the given slide IDs from source to destination."""
    dest_dir.mkdir(parents=True, exist_ok=True)

    for slide_id in slide_ids:
        src = source_dir / f"{slide_id}.xml"
        dst = dest_dir / f"{slide_id}.xml"

        if not src.exists():
            print(f"  [SKIP] Source not found: {src}")
            continue

        shutil.copy2(src, dst)
        print(f"  [COPY] {src.name} → {dst}")


def rename_groups_in_xml(xml_path: Path, renames: dict[str, str]) -> None:
    """
    Rename annotation groups in an ASAP XML file in-place.

    Affects:
      - <Annotation PartOfGroup="old"> → PartOfGroup="new"
      - <Group Name="old">            → Name="new"
    """
    ET.register_namespace("", "")  # keep no unwanted namespace prefixes
    tree = ET.parse(xml_path)
    root = tree.getroot()

    changed = 0

    # <Annotation ... PartOfGroup="old_name" ...>
    for annotation in root.iter("Annotation"):
        group = annotation.get("PartOfGroup")
        if group in renames:
            annotation.set("PartOfGroup", renames[group])
            changed += 1

    # <Group Name="old_name" ...>
    for group_elem in root.iter("Group"):
        name = group_elem.get("Name")
        if name in renames:
            group_elem.set("Name", renames[name])
            changed += 1

    if changed:
        # Preserve the XML declaration
        tree.write(xml_path, encoding="unicode", xml_declaration=True)
        print(f"  [RENAMED] {xml_path.name}: {changed} attribute(s) updated")
    else:
        print(f"  [NO CHANGE] {xml_path.name}: no matching group names found")


def main() -> None:
    print(f"Loading rename map from: {EXCEL_PATH}")
    rename_map = load_rename_map(EXCEL_PATH)
    print(f"  → {len(rename_map)} slide(s) with rename rules\n")

    print(f"Copying XMLs: {SOURCE_DIR} → {DEST_DIR}")
    copy_annotations(list(rename_map.keys()), SOURCE_DIR, DEST_DIR)
    print()

    print("Renaming annotation groups:")
    for slide_id, renames in rename_map.items():
        xml_path = DEST_DIR / f"{slide_id}.xml"
        if not xml_path.exists():
            print(f"  [SKIP] Not found after copy: {xml_path}")
            continue
        rename_groups_in_xml(xml_path, renames)


if __name__ == "__main__":
    main()
